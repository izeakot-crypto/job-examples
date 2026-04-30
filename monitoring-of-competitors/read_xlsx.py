import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook('C:/Users/izeak/Downloads/Competitor_Analysis_Template (2).xlsx')
ws = wb.active

print('=== HEADERS (Row 1) ===')
headers = [cell.value for cell in ws[1]]
for i, h in enumerate(headers, 1):
    print(f'{i}. {h}')

print()
print(f'Total rows: {ws.max_row}')
print()
print('=== SAMPLE DATA (Rows 2-4) ===')
for row_num in range(2, min(5, ws.max_row + 1)):
    print(f'--- Row {row_num} ---')
    for col, header in enumerate(headers, 1):
        val = ws.cell(row=row_num, column=col).value
        if val:
            val_str = str(val)[:150]
            print(f'  [{col}] {header}: {val_str}')
