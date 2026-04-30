import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook('C:/Users/izeak/Downloads/Competitor_Analysis_Template (2).xlsx')
ws = wb.active

print('=== АНАЛІЗ ЗАПОВНЕНОСТІ КОЛОНОК ===')
headers = [cell.value for cell in ws[1]]

# Count filled cells per column
column_stats = {}
for col, header in enumerate(headers, 1):
    if header is None:
        continue
    filled = 0
    empty = 0
    dash_only = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=col).value
        if val is None or str(val).strip() == '':
            empty += 1
        elif str(val).strip() == '-':
            dash_only += 1
        else:
            filled += 1
    column_stats[header] = {'filled': filled, 'empty': empty, 'dash_only': dash_only}

print(f'\nTotal data rows: {ws.max_row - 1}')
print()
print('КОЛОНКА | ЗАПОВНЕНО | ПУСТО | ТІЛЬКИ "-" | СТАТУС')
print('-' * 80)
for header, stats in column_stats.items():
    total = stats['filled'] + stats['empty'] + stats['dash_only']
    fill_pct = (stats['filled'] / total * 100) if total > 0 else 0
    status = 'OK' if fill_pct > 50 else 'PROBLEM' if fill_pct < 20 else 'NEEDS WORK'
    print(f'{header[:30]:30} | {stats["filled"]:9} | {stats["empty"]:5} | {stats["dash_only"]:10} | {status}')

print()
print('=== ПРОБЛЕМНІ ЗАПИСИ (Company = Unknown або URL = не вказано) ===')
problem_count = 0
for row in range(2, min(ws.max_row + 1, 20)):
    company = ws.cell(row=row, column=2).value
    url = ws.cell(row=row, column=3).value
    if company == 'Unknown' or url == 'не вказано':
        problem_count += 1
        print(f'Row {row}: Company="{company}", URL="{url}"')

# Count all Unknown
unknown_count = sum(1 for row in range(2, ws.max_row + 1) if ws.cell(row=row, column=2).value == 'Unknown')
print(f'\nTotal "Unknown" companies: {unknown_count}')
